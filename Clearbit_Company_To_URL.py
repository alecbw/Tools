# -*- coding: utf-8 -*-
from pprint import pprint
from openpyxl import Workbook , load_workbook
import requests


##############################################

IOFile = "InputFileString" + ".xlsx"

start_row = 2
end_row = 232
name_column = 1                 # The column from your input spreadsheet that the company names are in (e.g. Col A -> 1)

##############################################

wb1 = load_workbook(IOFile) # Importing our existing Workbook (Excel Sheet)
ws1 = wb1.active            # Detecting active sheet

wb2 = Workbook()            # Setting up new Workbook (Excel Sheet) to write to
ws2 = wb2.active            # Setting active sheet

ws2['A1'] = "Name"  # Setting up headers in the first row
ws2['B1'] = "Link"

Clearbit_Query_URL = "https://company.clearbit.com/v1/domains/find?" # Base query URL
CLEARBIT_KEY = "YOUR KEY GOES HERE"                                  # KEEP IT PRIVATE !!!!!

##############################################

def query_API(inputCompanyName, inputRowNumber):

    payload = {"name": inputCompanyName}                 # Setting up payload with our Company Names
    headers = {"Authorization":"Bearer " + CLEARBIT_KEY} # Setting Auth Header

    response = requests.request("GET", Clearbit_Query_URL, params=payload, headers=headers)
    JResponse = response.json()                          # Converting the response object to a JSON object

    pprint(response.json())                              # Typically commented out. Comment in if you'd like

    try:                                                 # Try / except keeps track of the unhelpfully different keys to the response object
        if JResponse['name']:                            # A successful 200 lookup should have a 'name' key
            ws2['A' + str(inputRowNumber)] = JResponse['name']   # Writing each variable to the Workbook
            ws2['B' + str(inputRowNumber)] = JResponse['domain']

    except KeyError:                # Thrown if a company name is not found
        print JResponse['error']    # Prints error text


def iterate_rows():                 # Helper function to iterate through imported spreadsheet

    for row in ws1.iter_rows(min_row=start_row, max_row=end_row, min_col=name_column, max_col=name_column):
        for cell in row:
            if not cell.value:                  # Detect and skip empty cells
                print "Skipping empty cell..."
                continue
            elif type(cell.value) == long:      # Weird error handling I had to add
                print "Skipping long type..."
                continue
            else:
                query_API(cell.value, cell.row)

if __name__ == '__main__':
    iterate_rows()
    wb2.save("CCTU Result for: " + IOFile) # Saving our file (as an xlsx filetype)



"""
Helpful resources:

https://dashboard.clearbit.com/api

https://blog.clearbit.com/company-name-to-domain-api/
https://clearbit.com/docs?python#authentication
https://github.com/clearbit/clearbit-python
https://github.com/dbugsy/clearbit-query/blob/master/lib/query.rb
https://github.com/clearbit/clearbit-python/blob/master/clearbit/tests.py#L18:99
"""
